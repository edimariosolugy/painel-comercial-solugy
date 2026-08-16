#!/usr/bin/env python3
"""
Gera data/funil.js a partir da planilha "Gestão e Controle de Orçamentos.xlsm".
Rode sempre que atualizar a planilha:
    python scripts/extract_funil.py "caminho/para/Gestão e Controle de Orçamentos.xlsm"
Requer: pip install openpyxl
"""
import json
import sys
import unicodedata
from datetime import date
from pathlib import Path

import openpyxl

RAIZ = Path(__file__).resolve().parent.parent
SAIDA = RAIZ / "data" / "funil.js"

MESES = ["JANEIRO", "FEVEREIRO", "MARÇO", "ABRIL", "MAIO", "JUNHO",
         "JULHO", "AGOSTO", "SETEMBRO", "OUTUBRO", "NOVEMBRO", "DEZEMBRO"]


def norm(s):
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s)).encode("ascii", "ignore").decode()
    return s.lower().replace("–", "-").replace("—", "-").strip()


def f(v):
    try:
        return round(float(v), 2)
    except (TypeError, ValueError):
        return 0.0


def main(caminho):
    wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)

    # ---- Prospecção (empresas e scores) ----
    ws = wb["Prospecção"]
    linhas = list(ws.iter_rows(values_only=True))
    prospects = []
    for r in linhas[1:]:
        if not r[0]:
            continue
        prospects.append({
            "empresa": str(r[0]).strip(),
            "cidade": str(r[5] or "").strip(),
            "uf": str(r[6] or "").strip(),
            "segmento": str(r[9] or "").strip(),
            "score": f(r[14]),
            "prioridade": str(r[15] or "").strip(),
            "ranking": r[16] if isinstance(r[16], (int, float)) else None,
            "solucao": str(r[17] or "").strip(),
            "gancho": str(r[18] or "").strip(),
        })
    prospects.sort(key=lambda p: (-p["score"], p["ranking"] or 999))

    # ---- Lançamentos (orçamentos mês a mês) ----
    ws = wb["Lançamentos"]
    mensal = {}   # "2026-01" -> {orcado, mc, qtd, aprovadoValor, aprovadoQtd}
    for r in ws.iter_rows(min_row=2, values_only=True):
        ano, mes = r[0], norm(r[1]).upper()
        if not ano or not mes:
            continue
        try:
            m = [norm(x).upper() for x in MESES].index(mes) + 1
        except ValueError:
            continue
        chave = f"{int(ano)}-{m:02d}"
        d = mensal.setdefault(chave, {"orcado": 0, "mc": 0, "qtd": 0, "aprovadoValor": 0, "aprovadoQtd": 0})
        d["orcado"] += f(r[7]); d["mc"] += f(r[6]); d["qtd"] += 1
        if norm(r[10]).startswith("aprovado"):
            d["aprovadoValor"] += f(r[7]); d["aprovadoQtd"] += 1
    for d in mensal.values():
        d["orcado"] = round(d["orcado"], 2); d["mc"] = round(d["mc"], 2)
        d["aprovadoValor"] = round(d["aprovadoValor"], 2)
        d["taxaAprov"] = round(100 * d["aprovadoQtd"] / d["qtd"], 1) if d["qtd"] else 0

    # ---- Fechamentos (MC por mês vs objetivo) ----
    ws = wb["Fechamentos"]
    fechamentos = []
    for r in ws.iter_rows(min_row=2, max_row=13, values_only=True):
        if not r[0]:
            continue
        fechamentos.append({"mes": str(r[0]), "mcVendas": f(r[1]), "mcOperacao": f(r[2]), "objetivo": f(r[4])})

    # ---- Tabelas mensais por segmento / origem ----
    def tabela(nome):
        ws = wb[nome]
        out = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r[0] or norm(r[0]) == "total":
                continue
            out.append({"nome": str(r[0]).strip(), "meses": [f(x) for x in r[1:13]], "total": f(r[13])})
        return out

    payload = {
        "geradoEm": date.today().isoformat(),
        "prospects": prospects,
        "mensal": dict(sorted(mensal.items())),
        "fechamentos": fechamentos,
        "orcadoPorSegmento": tabela("Segmentos dos Orçamentos"),
        "vendidoPorSegmento": tabela("Segmentos Vendas Fechadas"),
        "origemOrcamentos": tabela("Origem dos Orçamentos"),
    }
    # Não regrava (nem gera commit/deploy) se os dados não mudaram
    if SAIDA.exists():
        try:
            antigo = json.loads(SAIDA.read_text(encoding="utf-8").split("window.FUNIL = ", 1)[1].rstrip().rstrip(";"))
            antigo.pop("geradoEm", None)
            novo = dict(payload); novo.pop("geradoEm", None)
            if antigo == novo:
                print("Sem mudanças no funil — nada a gravar.")
                wb.close()
                return
        except Exception:
            pass
    SAIDA.write_text(
        "// Gerado por scripts/extract_funil.py — rode novamente após atualizar a planilha.\n"
        "window.FUNIL = " + json.dumps(payload, ensure_ascii=False) + ";\n",
        encoding="utf-8",
    )
    print(f"OK: {len(prospects)} empresas, {len(mensal)} meses de lançamentos → {SAIDA}")
    wb.close()


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "Gestão e Controle de Orçamentos.xlsm")
